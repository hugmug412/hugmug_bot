#!/usr/bin/env python3
"""
ربات حسابداری فروش هاگ ماگ - نسخه ۳
نسخه کامل با چند محصول، توضیحات، وضعیت سفارش
"""

import os
import re
import sqlite3
import jdatetime
from datetime import datetime, timedelta
from telegram import (
    Update, ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ContextTypes, ConversationHandler, CallbackQueryHandler
)
from reportlab.lib.pagesizes import A5
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ===== تنظیمات =====
BOT_TOKEN = os.environ.get("BOT_TOKEN")
if not BOT_TOKEN:
    raise ValueError("متغیر محیطی BOT_TOKEN تنظیم نشده!")
DB_FILE = "sales.db"
PROFIT_MARGIN = 0.30

# اطلاعات ثابت فروشگاه
STORE_NAME = "فروشگاه لوازم کادویی هاگ ماگ"
STORE_ADDRESS = "خوزستان، اندیمشک، خیابان سینا، جنب آزمایشگاه سینا"
STORE_PHONE = "09376923487"
LOGO_PATH = "logo.png"

# وضعیت‌های سفارش
STATUS_NEW      = "🟡 ثبت شده"
STATUS_READY    = "📦 آماده ارسال"
STATUS_SENT     = "🚚 ارسال شده"
ALL_STATUSES    = [STATUS_NEW, STATUS_READY, STATUS_SENT]

# States
WAITING_ITEMS   = 1
WAITING_NOTE    = 2
WAITING_LABEL   = 3

FONT_NAME = "Vazir"
FONT_REGISTERED = False

# ===== فونت فارسی =====
def register_persian_font():
    global FONT_REGISTERED
    if FONT_REGISTERED:
        return
    font_path = "Vazirmatn-Regular.ttf"
    if not os.path.exists(font_path):
        import urllib.request
        try:
            url = "https://github.com/rastikerdar/vazirmatn/raw/master/fonts/ttf/Vazirmatn-Regular.ttf"
            urllib.request.urlretrieve(url, font_path)
        except Exception:
            pass
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont(FONT_NAME, font_path))
        FONT_REGISTERED = True

def fa(text):
    reshaped = arabic_reshaper.reshape(str(text))
    return get_display(reshaped)

# ===== دیتابیس =====
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    # جدول سفارشات
    c.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            note TEXT DEFAULT "",
            status TEXT DEFAULT "🟡 ثبت شده",
            date TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            customer_name TEXT DEFAULT "",
            label_number INTEGER DEFAULT 0
        )
    ''')
    # اضافه کردن ستون‌های جدید به جداول قدیمی (برای upgrade بدون از دست دادن داده)
    for col in [("customer_name", "TEXT DEFAULT ''"), ("label_number", "INTEGER DEFAULT 0")]:
        try:
            conn.execute(f"ALTER TABLE orders ADD COLUMN {col[0]} {col[1]}")
        except Exception:
            pass
    # جدول اقلام هر سفارش
    c.execute('''
        CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER NOT NULL,
            product TEXT NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 1,
            amount REAL NOT NULL,
            FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
        )
    ''')
    # جدول شماره فاکتور
    c.execute('''
        CREATE TABLE IF NOT EXISTS label_counter (
            user_id INTEGER PRIMARY KEY,
            last_number INTEGER NOT NULL DEFAULT 0
        )
    ''')
    conn.commit()
    conn.close()

def create_order(user_id: int, items: list, note: str = "") -> int:
    """ثبت سفارش جدید با چند آیتم. items = [(product, qty, amount), ...]"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(
        "INSERT INTO orders (note, status, date, user_id, customer_name, label_number) VALUES (?, ?, ?, ?, ?, ?)",
        (note, STATUS_NEW, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), user_id, "", 0)
    )
    order_id = c.lastrowid
    for product, qty, amount in items:
        c.execute(
            "INSERT INTO order_items (order_id, product, quantity, amount) VALUES (?, ?, ?, ?)",
            (order_id, product, qty, amount)
        )
    conn.commit()
    conn.close()
    return order_id

def save_order_label_info(order_id: int, user_id: int, customer_name: str, label_number: int):
    """ذخیره نام مشتری و شماره فاکتور روی سفارش"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(
        "UPDATE orders SET customer_name=?, label_number=? WHERE id=? AND user_id=?",
        (customer_name, label_number, order_id, user_id)
    )
    conn.commit()
    conn.close()
    conn.close()
    return order_id

def get_order(order_id: int, user_id: int):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT id, note, status, date, customer_name, label_number FROM orders WHERE id=? AND user_id=?", (order_id, user_id))
    order = c.fetchone()
    if not order:
        conn.close()
        return None, []
    c.execute("SELECT id, product, quantity, amount FROM order_items WHERE order_id=?", (order_id,))
    items = c.fetchall()
    conn.close()
    return order, items

def get_orders_in_range(user_id: int, start: str, end: str):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(
        "SELECT id, note, status, date, customer_name, label_number FROM orders WHERE user_id=? AND date BETWEEN ? AND ? ORDER BY date DESC",
        (user_id, start, end)
    )
    orders = c.fetchall()
    result = []
    for o in orders:
        c.execute("SELECT product, quantity, amount FROM order_items WHERE order_id=?", (o[0],))
        items = c.fetchall()
        result.append((o, items))
    conn.close()
    return result

def get_status_counts(user_id: int):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    counts = {}
    for s in ALL_STATUSES:
        c.execute("SELECT COUNT(*) FROM orders WHERE user_id=? AND status=?", (user_id, s))
        counts[s] = c.fetchone()[0]
    conn.close()
    return counts

def update_order_status(order_id: int, user_id: int, new_status: str):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("UPDATE orders SET status=? WHERE id=? AND user_id=?", (new_status, order_id, user_id))
    conn.commit()
    conn.close()

def update_order_note(order_id: int, user_id: int, note: str):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("UPDATE orders SET note=? WHERE id=? AND user_id=?", (note, order_id, user_id))
    conn.commit()
    conn.close()

def delete_order(order_id: int, user_id: int):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
    c.execute("DELETE FROM orders WHERE id=? AND user_id=?", (order_id, user_id))
    conn.commit()
    deleted = c.rowcount > 0
    conn.close()
    return deleted

def get_product_summary(user_id: int, start: str, end: str):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(
        """SELECT oi.product, SUM(oi.amount * oi.quantity), SUM(oi.quantity)
           FROM order_items oi
           JOIN orders o ON oi.order_id = o.id
           WHERE o.user_id=? AND o.date BETWEEN ? AND ?
           GROUP BY oi.product ORDER BY 2 DESC""",
        (user_id, start, end)
    )
    rows = c.fetchall()
    conn.close()
    return rows

def get_next_label_number(user_id: int) -> int:
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("INSERT OR IGNORE INTO label_counter (user_id, last_number) VALUES (?, 0)", (user_id,))
    c.execute("UPDATE label_counter SET last_number = last_number + 1 WHERE user_id=?", (user_id,))
    conn.commit()
    c.execute("SELECT last_number FROM label_counter WHERE user_id=?", (user_id,))
    number = c.fetchone()[0]
    conn.close()
    return number

# ===== کیبورد =====
def main_keyboard():
    buttons = [
        [KeyboardButton("➕ ثبت فروش"), KeyboardButton("✏️ ویرایش/حذف")],
        [KeyboardButton("💰 امروز"), KeyboardButton("📊 گزارش هفتگی")],
        [KeyboardButton("📈 گزارش ماهانه"), KeyboardButton("🏆 برترین محصولات")],
        [KeyboardButton("📋 آخرین سفارشات"), KeyboardButton("🏷 فاکتور پستی")],
        [KeyboardButton("📥 خروجی اکسل"), KeyboardButton("📦 وضعیت سفارشات")],
    ]
    return ReplyKeyboardMarkup(buttons, resize_keyboard=True)

# ===== پارسر محصولات =====
def parse_items_text(text: str):
    """
    پارس کردن چند محصول از متن چند‌خطی.
    هر خط: نام/تعداد/قیمت  یا  نام، تعداد، قیمت
    """
    items = []
    errors = []
    for i, line in enumerate(text.strip().split("\n"), 1):
        line = line.strip()
        if not line:
            continue
        normalized = line.replace("،", "/").replace(",", "/")
        parts = [p.strip() for p in normalized.split("/")]
        if len(parts) != 3:
            errors.append(f"خط {i}: «{line}» (باید ۳ بخش داشته باشه)")
            continue
        product, qty_str, amount_str = parts
        try:
            qty = int(qty_str.replace(" ", ""))
            amount = float(amount_str.replace(" ", ""))
            if qty <= 0 or amount < 0 or not product:
                raise ValueError
            items.append((product, qty, amount))
        except ValueError:
            errors.append(f"خط {i}: «{line}» (تعداد یا قیمت نامعتبر)")
    return items, errors

def format_number(n):
    return f"{n:,.0f}"

def order_total(items):
    return sum(a * q for (_, q, a) in items)

# ===== تاریخ شمسی =====
def to_jalali(value, fmt="%Y/%m/%d"):
    """
    تبدیل تاریخ میلادی به شمسی برای نمایش.
    value می‌تونه datetime باشه یا رشته‌ی ذخیره‌شده در دیتابیس ('YYYY-MM-DD HH:MM:SS' یا 'YYYY-MM-DD').
    """
    if isinstance(value, str):
        v = value.split(".")[0].strip()
        if " " in v:
            value = datetime.strptime(v, "%Y-%m-%d %H:%M:%S")
        else:
            value = datetime.strptime(v, "%Y-%m-%d")
    j = jdatetime.datetime.fromgregorian(datetime=value)
    return j.strftime(fmt)

def get_jalali_month_bounds(offset: int = 0):
    """
    محدوده‌ی یک ماه شمسی رو برمی‌گردونه.
    offset=0 یعنی ماه جاری، offset=1 یعنی یک ماه قبل، offset=2 دو ماه قبل و ...
    خروجی: (start_datetime, end_datetime, برچسب 'YYYY/MM')
    """
    today_j = jdatetime.date.today()
    total = today_j.year * 12 + (today_j.month - 1) - offset
    y, m0 = divmod(total, 12)
    m = m0 + 1
    start_j = jdatetime.date(y, m, 1)
    if m == 12:
        next_start_j = jdatetime.date(y + 1, 1, 1)
    else:
        next_start_j = jdatetime.date(y, m + 1, 1)
    end_j = next_start_j - timedelta(days=1)

    start_g = datetime.combine(start_j.togregorian(), datetime.min.time())
    if offset == 0:
        end_g = datetime.now()
    else:
        end_g = datetime.combine(end_j.togregorian(), datetime.max.time())

    label = start_j.strftime("%Y/%m")
    return start_g, end_g, label

# ===== هندلرها =====
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.effective_user.first_name
    await update.message.reply_text(
        f"سلام {name}! 👋\n\n"
        "📦 برای ثبت فروش، هر محصول رو تو یه خط بنویس:\n"
        "نام محصول/تعداد/قیمت واحد\n\n"
        "مثال (یک محصول):\n"
        "تراول ماگ دیلر/۱/۱۵۰۰۰۰\n\n"
        "مثال (چند محصول):\n"
        "تراول ماگ دیلر/۱/۱۵۰۰۰۰\n"
        "ماگ سرامیکی/۲/۹۰۰۰۰\n\n"
        "یا از منوی پایین استفاده کن:",
        reply_markup=main_keyboard()
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = update.effective_user.id

    # --- منتظر توضیحات سفارش ---
    if context.user_data.get('awaiting_note'):
        await process_note(update, context)
        return ConversationHandler.END

    # --- منتظر آدرس فاکتور ---
    if context.user_data.get('awaiting_order_address'):
        await process_order_address(update, context)
        return ConversationHandler.END

    # --- منتظر اسم مشتری ---
    if context.user_data.get('awaiting_customer_name'):
        await process_customer_name(update, context)
        return ConversationHandler.END

    # --- دکمه‌های منو ---
    if text == "➕ ثبت فروش":
        await update.message.reply_text(
            "📦 محصولات رو بفرست (هر خط یه محصول):\n"
            "نام/تعداد/قیمت\n\n"
            "مثال:\n"
            "تراول ماگ دیلر/۱/۱۵۰۰۰۰\n"
            "ماگ سرامیکی/۲/۹۰۰۰۰\n\n"
            "(برای لغو /cancel بزن)"
        )
        return WAITING_ITEMS

    elif text == "✏️ ویرایش/حذف":
        await send_edit_menu(update, context)

    elif text == "💰 امروز":
        await send_today_report(update, context)

    elif text == "📊 گزارش هفتگی":
        await send_weekly_report(update, context)

    elif text == "📈 گزارش ماهانه":
        await send_monthly_report(update, context)

    elif text == "🏆 برترین محصولات":
        await send_top_products(update, context)

    elif text == "📋 آخرین سفارشات":
        await send_recent_orders(update, context)

    elif text == "🏷 فاکتور پستی":
        await update.message.reply_text(
            "📮 مشخصات گیرنده رو بفرست (هر چی هست، همونجوری پیست کن):\n\n"
            "(برای لغو /cancel بزن)"
        )
        context.user_data['awaiting_order_address'] = True
        context.user_data['standalone_label'] = True

    elif text == "📥 خروجی اکسل":
        await send_excel_export(update, context)

    elif text == "📦 وضعیت سفارشات":
        await send_status_overview(update, context)

    else:
        # ثبت سریع بدون دکمه
        items, errors = parse_items_text(text)
        if items and not errors:
            await quick_register(update, context, items)
        # اگه نه، نادیده بگیر

    return ConversationHandler.END

async def get_items(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """دریافت آیتم‌ها از ConversationHandler state"""
    items, errors = parse_items_text(update.message.text)

    if errors:
        msg = "❌ برخی خطوط نامعتبر بودن:\n" + "\n".join(errors)
        msg += "\n\nدوباره امتحان کن یا /cancel بزن."
        await update.message.reply_text(msg)
        return WAITING_ITEMS

    if not items:
        await update.message.reply_text("❌ هیچ محصولی پارس نشد. دوباره امتحان کن.")
        return WAITING_ITEMS

    await register_and_ask_address(update, context, items)
    return ConversationHandler.END

async def quick_register(update: Update, context: ContextTypes.DEFAULT_TYPE, items: list):
    """ثبت سریع بدون دکمه"""
    await register_and_ask_address(update, context, items)

async def register_and_ask_address(update: Update, context: ContextTypes.DEFAULT_TYPE, items: list):
    """ثبت سفارش (بدون توضیحات فعلاً) و پرسیدن آدرس"""
    user_id = update.effective_user.id
    order_id = create_order(user_id, items, note="")
    total = order_total(items)
    summary = "\n".join(f"• {p} × {q}" for p, q, a in items)

    context.user_data['last_order_id'] = order_id

    keyboard = [[
        InlineKeyboardButton("🏷 بله، آدرس می‌فرستم", callback_data=f"addaddr_yes_{order_id}"),
        InlineKeyboardButton("❌ نه", callback_data="addaddr_no"),
    ]]
    await update.message.reply_text(
        f"🎉 سفارش #{order_id} ثبت شد!\n\n"
        f"{summary}\n"
        f"💰 جمع: {format_number(total)} تومان\n"
        f"🟡 وضعیت: ثبت شده\n\n"
        "📮 آدرس پستی هم می‌خوای ثبت کنی؟",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("❌ لغو شد.", reply_markup=main_keyboard())
    return ConversationHandler.END

# ===== ویرایش / حذف / وضعیت =====
async def send_edit_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    end = datetime.now()
    start = end - timedelta(days=30)
    orders_data = get_orders_in_range(user_id,
        start.strftime("%Y-%m-%d 00:00:00"),
        end.strftime("%Y-%m-%d 23:59:59")
    )[:10]

    if not orders_data:
        await update.message.reply_text("📭 سفارشی برای ویرایش وجود نداره.")
        return

    for (oid, note, status, date, customer_name, label_num), items in orders_data:
        d = to_jalali(date, "%m/%d")
        total = sum(a * q for _, q, a in items)
        products_str = "، ".join(f"{p}×{q}" for p, q, a in items[:2])
        if len(items) > 2:
            products_str += f" (+{len(items)-2})"
        keyboard = [[
            InlineKeyboardButton("🔄 وضعیت", callback_data=f"setstatus_{oid}"),
            InlineKeyboardButton("📝 توضیحات", callback_data=f"editnote_{oid}"),
            InlineKeyboardButton("🗑 حذف", callback_data=f"del_{oid}"),
        ]]
        await update.message.reply_text(
            f"#{oid} | {d} | {status}\n{products_str}\n💰 {format_number(total)} ت",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

async def handle_callbacks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = update.effective_user.id
    data = query.data

    # --- آدرس پستی ---
    if data.startswith("addaddr_yes_"):
        order_id = int(data.split("_")[2])
        context.user_data['label_order_id'] = order_id
        await query.edit_message_text(
            "📮 مشخصات گیرنده رو بفرست (همونجوری پیست کن):\n(برای لغو /cancel)"
        )
        context.user_data['awaiting_order_address'] = True

    elif data == "addaddr_no":
        context.user_data.pop('last_order_id', None)
        await query.edit_message_text("✅ باشه، سفارش ثبت موند.")

    # --- وضعیت سفارش ---
    elif data.startswith("setstatus_"):
        order_id = int(data.split("_")[1])
        keyboard = [[InlineKeyboardButton(s, callback_data=f"dostatus_{order_id}_{i}")]
                    for i, s in enumerate(ALL_STATUSES)]
        await query.edit_message_text("🔄 وضعیت جدید رو انتخاب کن:", reply_markup=InlineKeyboardMarkup(keyboard))

    elif data.startswith("dostatus_"):
        _, order_id_str, idx_str = data.split("_")
        order_id = int(order_id_str)
        new_status = ALL_STATUSES[int(idx_str)]
        update_order_status(order_id, user_id, new_status)
        await query.edit_message_text(f"✅ وضعیت سفارش #{order_id} شد: {new_status}")

    # --- ویرایش توضیحات ---
    elif data.startswith("editnote_"):
        order_id = int(data.split("_")[1])
        context.user_data['editing_note_order_id'] = order_id
        context.user_data['awaiting_note_edit'] = True
        await query.edit_message_text(
            f"📝 توضیحات جدید برای سفارش #{order_id} رو بفرست:\n(یا /skip برای پاک کردن)"
        )

    # --- حذف ---
    elif data.startswith("del_"):
        order_id = int(data.split("_")[1])
        keyboard = [[
            InlineKeyboardButton("✅ بله حذف کن", callback_data=f"confirmdel_{order_id}"),
            InlineKeyboardButton("❌ لغو", callback_data="canceldel"),
        ]]
        order, items = get_order(order_id, user_id)
        if order:
            total = sum(a * q for _, _, q, a in items)
            await query.edit_message_text(
                f"⚠️ سفارش #{order_id} حذف بشه؟\n💰 {format_number(total)} تومان",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )

    elif data.startswith("confirmdel_"):
        order_id = int(data.split("_")[1])
        delete_order(order_id, user_id)
        await query.edit_message_text(f"✅ سفارش #{order_id} حذف شد.")

    elif data == "canceldel":
        await query.edit_message_text("❌ حذف لغو شد.")

    # --- روش ارسال ---
    elif data.startswith("shipvia_"):
        method = data.split("_")[1]
        method_names = {"tipax": "تیپاکس", "post": "پست", "chapar": "چاپار"}
        method_fa = method_names.get(method, method)
        order = context.user_data.get('pending_label_order')
        if not order:
            await query.edit_message_text("❌ اطلاعات پیدا نشد.")
            return
        if method == "post":
            order['shipping_method'] = method_fa
            context.user_data['pending_label_order'] = order
            keyboard = [[
                InlineKeyboardButton("💰 پیش‌کرایه", callback_data="postpay_pre"),
                InlineKeyboardButton("💵 پس‌کرایه", callback_data="postpay_post"),
            ]]
            await query.edit_message_text("📮 نوع کرایه پست چیه؟", reply_markup=InlineKeyboardMarkup(keyboard))
            return
        await finalize_label(update, context, order, method_fa, user_id)

    elif data.startswith("postpay_"):
        pay_fa = "پیش‌کرایه" if data.split("_")[1] == "pre" else "پس‌کرایه"
        order = context.user_data.get('pending_label_order')
        if not order:
            await query.edit_message_text("❌ اطلاعات پیدا نشد.")
            return
        method_fa = f"{order.get('shipping_method', 'پست')} ({pay_fa})"
        await finalize_label(update, context, order, method_fa, user_id)

    # --- ناوبری گزارش ماهانه ---
    elif data.startswith("monthrep_"):
        offset = int(data.split("_")[1])
        await send_monthly_report(update, context, offset=offset, edit=True)

    # --- ناوبری خروجی اکسل ---
    elif data.startswith("excelrep_"):
        offset = int(data.split("_")[1])
        await send_excel_export(update, context, offset=offset, from_callback=True)

async def handle_note_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ویرایش توضیحات از منوی ویرایش"""
    if not context.user_data.get('awaiting_note_edit'):
        return
    text = update.message.text.strip()
    user_id = update.effective_user.id
    order_id = context.user_data.pop('editing_note_order_id', None)
    context.user_data.pop('awaiting_note_edit', None)
    if not order_id:
        return
    note = "" if text in ["/skip", "ندارم", "-"] else text
    update_order_note(order_id, user_id, note)
    await update.message.reply_text(f"✅ توضیحات سفارش #{order_id} آپدیت شد.", reply_markup=main_keyboard())

# ===== فلوی فاکتور پستی =====
async def process_order_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    address_text = update.message.text.strip()
    context.user_data['pending_label_order'] = {
        'recipient_info': address_text,
        'order_id': context.user_data.get('label_order_id'),
        'standalone': context.user_data.get('standalone_label', False),
    }
    context.user_data.pop('awaiting_order_address', None)
    context.user_data.pop('label_order_id', None)
    context.user_data.pop('standalone_label', None)
    context.user_data['awaiting_customer_name'] = True
    await update.message.reply_text("👤 اسم مشتری چی باشه؟ (برای نام‌گذاری فایل)")

async def process_customer_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    customer_name = update.message.text.strip()
    order = context.user_data.get('pending_label_order')
    if not order:
        await update.message.reply_text("❌ سفارشی پیدا نشد.", reply_markup=main_keyboard())
        context.user_data.clear()
        return
    order['customer_name'] = customer_name
    context.user_data['pending_label_order'] = order
    context.user_data.pop('awaiting_customer_name', None)
    # بعد از اسم مشتری، توضیحات سفارش رو بپرس
    context.user_data['awaiting_note'] = True
    await update.message.reply_text(
        "📝 توضیحات سفارش داری؟ (مثل: کادو شود، رنگ مشکی)\n"
        "اگه نداری بنویس /skip"
    )

async def process_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """توضیحات بعد از آدرس و نام مشتری — ذخیره میشه و بعد روش ارسال پرسیده می‌شه"""
    text = update.message.text.strip()
    user_id = update.effective_user.id
    order = context.user_data.get('pending_label_order')

    if not order:
        await update.message.reply_text("❌ سفارشی پیدا نشد.", reply_markup=main_keyboard())
        context.user_data.clear()
        return

    note = "" if text in ["/skip", "ندارم", "-"] else text
    order['note'] = note
    context.user_data['pending_label_order'] = order
    context.user_data.pop('awaiting_note', None)

    # ذخیره توضیحات روی سفارش در دیتابیس
    order_id = order.get('order_id')
    if order_id and note:
        update_order_note(order_id, user_id, note)

    keyboard = [[
        InlineKeyboardButton("📦 تیپاکس", callback_data="shipvia_tipax"),
        InlineKeyboardButton("📮 پست", callback_data="shipvia_post"),
        InlineKeyboardButton("🚚 چاپار", callback_data="shipvia_chapar"),
    ]]
    await update.message.reply_text("🚚 روش ارسال چیه؟", reply_markup=InlineKeyboardMarkup(keyboard))

async def finalize_label(update, context, order: dict, method_fa: str, user_id: int):
    query = update.callback_query
    order_id = order.get('order_id')
    items_for_label = []
    note = ""

    if order_id:
        db_order, db_items = get_order(order_id, user_id)
        items_for_label = [(p, q, a) for _, p, q, a in db_items]
        if db_order:
            note = db_order[1]  # note field

    try:
        serial = get_next_label_number(user_id)
        customer_name = order.get('customer_name', '').strip()

        # ذخیره نام مشتری و شماره فاکتور روی سفارش
        if order_id:
            save_order_label_info(order_id, user_id, customer_name, serial)

        pdf_path = create_shipping_label(
            order['recipient_info'],
            items_for_label,
            method_fa,
            note,
            order_id=order_id,
            label_number=serial
        )
        safe_name = re.sub(r'[^\w\u0600-\u06FF]+', '_', customer_name).strip('_') or "بدون_نام"
        file_name = f"فاکتور_{serial:03d}_{safe_name}.pdf"

        with open(pdf_path, "rb") as f:
            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=f,
                filename=file_name,
                caption=f"🏷 فاکتور ({method_fa}) | {customer_name}\n📄 {file_name}"
            )
        await query.edit_message_text(f"✅ فاکتور {file_name} ارسال شد.")
    except Exception as e:
        await query.edit_message_text(f"❌ خطا: {e}")

    context.user_data.pop('pending_label_order', None)

# ===== گزارش‌ها =====
async def send_today_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    today = datetime.now().strftime("%Y-%m-%d")
    orders_data = get_orders_in_range(user_id, today + " 00:00:00", today + " 23:59:59")
    if not orders_data:
        await update.message.reply_text("📭 امروز هیچ سفارشی ثبت نشده.")
        return
    total = sum(order_total(items) for _, items in orders_data)
    total_items = sum(sum(q for _, q, _ in items) for _, items in orders_data)
    text = f"☀️ گزارش امروز\n━━━━━━━━━━━━━━━\n"
    for (oid, note, status, date, customer_name, label_num), items in orders_data:
        t = date.split(" ")[1][:5]
        text += f"🕐 {t} | #{oid} | {status}\n"
        for p, q, a in items:
            text += f"   • {p} × {q} = {format_number(a*q)} ت\n"
    text += f"━━━━━━━━━━━━━━━\n🛒 {len(orders_data)} سفارش | {total_items} آیتم\n💰 {format_number(total)} تومان"
    await update.message.reply_text(text)

async def send_weekly_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    end = datetime.now()
    start = end - timedelta(days=7)
    orders_data = get_orders_in_range(user_id, start.strftime("%Y-%m-%d 00:00:00"), end.strftime("%Y-%m-%d 23:59:59"))
    if not orders_data:
        await update.message.reply_text("📭 در هفته گذشته سفارشی ثبت نشده.")
        return
    total = sum(order_total(items) for _, items in orders_data)
    total_items = sum(sum(q for _, q, _ in items) for _, items in orders_data)
    text = f"📊 گزارش هفتگی\n{to_jalali(start)} تا {to_jalali(end)}\n━━━━━━━━━━━━━━━\n"
    text += f"🛒 {len(orders_data)} سفارش | {total_items} آیتم\n"
    text += f"💰 درآمد کل: {format_number(total)} تومان\n"
    text += f"📌 میانگین روزانه: {format_number(total/7)} تومان\n"
    products = get_product_summary(user_id, start.strftime("%Y-%m-%d 00:00:00"), end.strftime("%Y-%m-%d 23:59:59"))
    if products:
        text += "\n🏆 برترین محصولات:\n"
        for i, (p, tot, qty) in enumerate(products[:5], 1):
            text += f"{i}. {p}: {qty} عدد | {format_number(tot)} ت\n"
    await update.message.reply_text(text)

async def send_monthly_report(update: Update, context: ContextTypes.DEFAULT_TYPE, offset: int = 0, edit: bool = False):
    user_id = update.effective_user.id
    start, end, label = get_jalali_month_bounds(offset)
    orders_data = get_orders_in_range(user_id, start.strftime("%Y-%m-%d 00:00:00"), end.strftime("%Y-%m-%d 23:59:59"))

    nav_row = [InlineKeyboardButton("◀️ ماه قبل", callback_data=f"monthrep_{offset+1}")]
    if offset > 0:
        nav_row.append(InlineKeyboardButton("ماه بعد ▶️", callback_data=f"monthrep_{offset-1}"))
    reply_markup = InlineKeyboardMarkup([nav_row])

    if not orders_data:
        text = f"📭 در {label} سفارشی ثبت نشده."
        if edit:
            await update.callback_query.edit_message_text(text, reply_markup=reply_markup)
        else:
            await update.message.reply_text(text, reply_markup=reply_markup)
        return

    total = sum(order_total(items) for _, items in orders_data)
    total_items = sum(sum(q for _, q, _ in items) for _, items in orders_data)
    days = (end.date() - start.date()).days + 1
    profit = total * PROFIT_MARGIN
    text = f"📈 گزارش ماهانه | {label}\n━━━━━━━━━━━━━━━\n"
    text += f"🛒 {len(orders_data)} سفارش | {total_items} آیتم\n"
    text += f"💰 درآمد کل: {format_number(total)} تومان\n"
    text += f"📌 میانگین روزانه: {format_number(total/days)} تومان\n"
    if offset == 0:
        text += f"📆 پیش‌بینی ماه: {format_number(total/days*30)} تومان\n"
    text += f"💵 سود تقریبی (۳۰٪): {format_number(profit)} تومان\n"
    products = get_product_summary(user_id, start.strftime("%Y-%m-%d 00:00:00"), end.strftime("%Y-%m-%d 23:59:59"))
    if products:
        text += "\n🏆 برترین محصولات:\n"
        for i, (p, tot, qty) in enumerate(products, 1):
            pct = tot / total * 100 if total else 0
            text += f"{i}. {p}: {qty} عدد | {format_number(tot)} ت ({pct:.0f}%)\n"

    if edit:
        await update.callback_query.edit_message_text(text, reply_markup=reply_markup)
    else:
        await update.message.reply_text(text, reply_markup=reply_markup)

async def send_top_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    end = datetime.now()
    start = end - timedelta(days=30)
    products = get_product_summary(user_id, start.strftime("%Y-%m-%d 00:00:00"), end.strftime("%Y-%m-%d 23:59:59"))
    if not products:
        await update.message.reply_text("📭 داده‌ای برای نمایش وجود نداره.")
        return
    total_all = sum(t for _, t, _ in products)
    medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
    text = "🏆 برترین محصولات (۳۰ روز اخیر)\n━━━━━━━━━━━━━━━\n"
    for i, (p, tot, qty) in enumerate(products):
        medal = medals[i] if i < len(medals) else f"{i+1}."
        pct = tot / total_all * 100 if total_all else 0
        bar = "█" * int(pct / 10) + "░" * (10 - int(pct / 10))
        text += f"{medal} {p}\n   {bar} {pct:.0f}%\n   {qty} عدد | {format_number(tot)} ت\n\n"
    await update.message.reply_text(text)

async def send_recent_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    end = datetime.now()
    start = end - timedelta(days=30)
    orders_data = get_orders_in_range(user_id, start.strftime("%Y-%m-%d 00:00:00"), end.strftime("%Y-%m-%d 23:59:59"))[:10]
    if not orders_data:
        await update.message.reply_text("📭 سفارشی ثبت نشده.")
        return
    text = "📋 آخرین سفارشات\n━━━━━━━━━━━━━━━\n"
    for (oid, note, status, date, customer_name, label_num), items in orders_data:
        d = to_jalali(date)
        t = date.split(" ")[1][:5]
        total = order_total(items)
        text += f"#{oid} | {d} {t} | {status}\n"
        for p, q, a in items:
            text += f"   • {p} × {q} = {format_number(a*q)} ت\n"
        if note:
            text += f"   📝 {note}\n"
        text += f"   💰 جمع: {format_number(total)} ت\n\n"
    await update.message.reply_text(text)

async def send_status_overview(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    end = datetime.now()
    start = end - timedelta(days=60)
    orders_data = get_orders_in_range(user_id,
        start.strftime("%Y-%m-%d 00:00:00"),
        end.strftime("%Y-%m-%d 23:59:59")
    )

    if not orders_data:
        await update.message.reply_text("📭 سفارشی در ۶۰ روز اخیر وجود نداره.")
        return

    # خلاصه وضعیت‌ها
    counts = get_status_counts(user_id)
    summary = "📦 وضعیت سفارشات\n━━━━━━━━━━━━━━━\n"
    for status, count in counts.items():
        summary += f"{status}: {count}\n"
    await update.message.reply_text(summary)

    # لیست سفارشات با دکمه تغییر وضعیت
    for (oid, note, status, date, customer_name, label_num), items in orders_data:
        d = to_jalali(date, "%m/%d")
        products_str = "، ".join(f"{p}×{q}" for p, q, a in items[:2])
        if len(items) > 2:
            products_str += f" (+{len(items)-2})"

        label_str = f"فاکتور_{label_num:03d}" if label_num else "بدون فاکتور"
        name_str = customer_name if customer_name else "—"

        keyboard = [[InlineKeyboardButton(s, callback_data=f"dostatus_{oid}_{i}")]
                    for i, s in enumerate(ALL_STATUSES)]

        await update.message.reply_text(
            f"#{oid} | {label_str} | {name_str}\n"
            f"📅 {d} | {status}\n"
            f"🛍 {products_str}",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

# ===== خروجی اکسل =====
async def send_excel_export(update: Update, context: ContextTypes.DEFAULT_TYPE, offset: int = 0, from_callback: bool = False):
    user_id = update.effective_user.id
    start, end, label = get_jalali_month_bounds(offset)
    orders_data = get_orders_in_range(user_id, start.strftime("%Y-%m-%d 00:00:00"), end.strftime("%Y-%m-%d 23:59:59"))

    nav_row = [InlineKeyboardButton("◀️ ماه قبل", callback_data=f"excelrep_{offset+1}")]
    if offset > 0:
        nav_row.append(InlineKeyboardButton("ماه بعد ▶️", callback_data=f"excelrep_{offset-1}"))
    reply_markup = InlineKeyboardMarkup([nav_row])

    target = update.callback_query.message if from_callback else update.message

    if not orders_data:
        await target.reply_text(f"📭 در {label} سفارشی برای خروجی گرفتن وجود نداره.", reply_markup=reply_markup)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "سفارشات ماه"
    ws.sheet_view.rightToLeft = True

    headers = ["شماره سفارش", "تاریخ", "نام محصول", "تعداد", "قیمت واحد", "مبلغ کل", "توضیحات", "وضعیت"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    total_all = 0
    for (oid, note, status, date, customer_name, label_num), items in sorted(orders_data, key=lambda x: x[0][3]):
        d = to_jalali(date)
        for p, q, a in items:
            row_total = a * q
            total_all += row_total
            ws.append([oid, d, p, q, a, row_total, note, status])

    ws.append(["", "", "", "", "", total_all, "", ""])
    last = ws.max_row
    ws[f"F{last}"].font = Font(bold=True)

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 16

    file_label = label.replace("/", "_")
    filename = f"/tmp/orders_{file_label}.xlsx"
    wb.save(filename)
    with open(filename, "rb") as f:
        await target.reply_document(
            document=f,
            filename=f"سفارشات_{file_label}.xlsx",
            caption=f"📥 خروجی اکسل {label} ✅",
            reply_markup=reply_markup
        )

# ===== فاکتور PDF =====
def create_shipping_label(recipient_info: str, items: list, shipping_method: str = "", note: str = "", order_id: int = 0, label_number: int = 0):
    register_persian_font()
    use_font = FONT_NAME if FONT_REGISTERED else "Helvetica"
    filename = f"/tmp/label_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    width, height = A5
    c = canvas.Canvas(filename, pagesize=A5)
    margin = 10 * mm
    y = height - margin

    def draw_rtl(text_str, font_size, y_pos):
        c.setFont(use_font, font_size)
        display = fa(text_str) if FONT_REGISTERED else text_str
        c.drawRightString(width - margin, y_pos, display)

    def wrap_rtl(text_str, max_chars, font_size, start_y):
        cur_y = start_y
        for raw_line in text_str.split("\n"):
            raw_line = raw_line.strip()
            if not raw_line:
                continue
            words = raw_line.split(" ")
            current = ""
            for word in words:
                test = (current + " " + word).strip()
                if len(test) > max_chars:
                    if current:
                        draw_rtl(current, font_size, cur_y)
                        cur_y -= 6 * mm
                    current = word
                else:
                    current = test
            if current:
                draw_rtl(current, font_size, cur_y)
                cur_y -= 6 * mm
        return cur_y

    # کادر
    c.setLineWidth(1.5)
    c.rect(5*mm, 5*mm, width - 10*mm, height - 10*mm)

    # لوگو
    logo_size = 18 * mm
    if os.path.exists(LOGO_PATH):
        try:
            logo = ImageReader(LOGO_PATH)
            c.drawImage(logo, width - margin - logo_size, height - margin - logo_size,
                        width=logo_size, height=logo_size, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    # عنوان
    y -= (logo_size + 3*mm)
    c.setFont(use_font, 15)
    title = fa("فاکتور ارسال مرسوله") if FONT_REGISTERED else "فاکتور ارسال مرسوله"
    c.drawCentredString(width / 2, y, title)
    y -= 6*mm
    c.setFont(use_font, 9)
    ref_parts = []
    if label_number:
        ref_parts.append(f"فاکتور: {label_number:03d}")
    if order_id:
        ref_parts.append(f"سفارش: #{order_id}")
    if ref_parts:
        ref = fa("  |  ".join(ref_parts)) if FONT_REGISTERED else "  |  ".join(ref_parts)
        c.drawCentredString(width / 2, y, ref)
        y -= 5*mm
    if shipping_method:
        c.setFont(use_font, 10)
        m = fa(f"روش ارسال: {shipping_method}") if FONT_REGISTERED else f"روش ارسال: {shipping_method}"
        c.drawCentredString(width / 2, y, m)
        y -= 5*mm
    y -= 3*mm

    c.setLineWidth(0.5)
    c.line(margin, y, width - margin, y)
    y -= 7*mm

    # فرستنده
    draw_rtl("فرستنده:", 11, y); y -= 6*mm
    draw_rtl(STORE_NAME, 11, y); y -= 6*mm
    y = wrap_rtl(STORE_ADDRESS, 42, 9, y)
    draw_rtl(f"تلفن: {STORE_PHONE}", 9, y); y -= 6*mm
    y -= 2*mm
    c.line(margin, y, width - margin, y)
    y -= 7*mm

    # گیرنده
    draw_rtl("گیرنده:", 11, y); y -= 6*mm
    y = wrap_rtl(recipient_info, 42, 10, y)
    y -= 2*mm
    c.line(margin, y, width - margin, y)
    y -= 7*mm

    # اقلام (بدون مبلغ)
    draw_rtl("اقلام سفارش:", 11, y); y -= 6*mm
    if items:
        for p, q, a in items:
            draw_rtl(f"• {p}  ×  {q}", 10, y); y -= 6*mm
    if note:
        y -= 2*mm
        draw_rtl(f"📝 {note}", 9, y); y -= 5*mm

    c.setFont(use_font, 8)
    footer_date = fa(to_jalali(datetime.now(), "%Y/%m/%d %H:%M")) if FONT_REGISTERED else to_jalali(datetime.now(), "%Y/%m/%d %H:%M")
    c.drawCentredString(width / 2, 10*mm, footer_date)
    c.save()
    return filename

# ===== اجرا =====
def main():
    init_db()
    register_persian_font()
    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message)],
        states={
            WAITING_ITEMS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_items)],
        },
        fallbacks=[CommandHandler("cancel", cancel), CommandHandler("skip", cancel)],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("skip", lambda u, c: process_note(u, c)))
    app.add_handler(conv)
    app.add_handler(CallbackQueryHandler(handle_callbacks))
    app.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND,
        lambda u, c: handle_note_edit(u, c) if c.user_data.get('awaiting_note_edit') else None
    ))

    print("✅ ربات هاگ ماگ نسخه ۳ شروع به کار کرد...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
